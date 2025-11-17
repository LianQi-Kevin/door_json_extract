import base64
import json
import os
import threading
from concurrent.futures import ThreadPoolExecutor, wait, ALL_COMPLETED
from typing import Optional
from dataclasses import dataclass

import pymupdf
import openpyxl
from openpyxl.styles.builtins import styles

from requests_tools import post_with_retry


# PaddleOCR-VL 配置
PADDLE_OCR_VL_DEPLOY_URL: str = r"http://your-paddle-ocr-vl-deploy-url:port"

# CHAT-GLM 配置
GLM_API_KEY: str = r"your-api-key"
CHAT_COMPLETIONS_URL: str = "https://open.bigmodel.cn/api/paas/v4/chat/completions"
MODEL_NAME: str = "glm-4-plus"


# cache data
@dataclass
class cacheData:
    pdf_path: str
    base64: Optional[str] = None
    markdown_text: Optional[str] = None
    extracted_json: Optional[dict] = None


CACHE_DATA_DICT: dict[str, cacheData] = {}

# thread lock
LOCK_1 = threading.Lock()


def paddle_ocr_vl(detection: str) -> str:
    global PADDLE_OCR_VL_DEPLOY_URL
    response = post_with_retry(
        url=f"{PADDLE_OCR_VL_DEPLOY_URL}/layout-parsing",
        headers={"Content-Type": "application/json"},
        data=json.dumps({
            "file": detection,
            "fileType": 1,
            "visualize": False,
            "prettifyMarkdown": False,
            "useLayoutDetection": True,
            "promptLabel": "table"
        }, ensure_ascii=False),
    )
    response.raise_for_status()
    data = response.json()
    try:
        return data["result"]["layoutParsingResults"][0]["markdown"]["text"]
    except (KeyError, IndexError, TypeError) as e:
        raise RuntimeError(f"PaddleOCR-VL返回结构异常: {data}") from e


def big_model_completions(markdown_text: str) -> dict:
    global GLM_API_KEY, CHAT_COMPLETIONS_URL, MODEL_NAME
    # 定义要提取的JSON schema
    json_schema: dict = {
        "门编号": "",
        "门型": "",
        "洞口尺寸": "",
        "构件尺寸": "",
        "门框材质": "",
        "门扇材质": "",
        "门槛材质": "",
        "防火门芯": "",
        "玻璃": "",
        "门框密封条": "",
        "门扇密封条": "",
        "五金配置组名称": "",
        "五金配置": [
            {"名称": "", "品牌": "", "型号": "", "数量": 0}
        ],
        "饰面颜色": {"推门侧": "", "拉门侧": ""}
    }

    # 定义系统提示模板
    system_template: str = f"""
    你是信息抽取引擎。任务：从输入文本中的表格中抽取下列字段，并严格以“唯一一个 JSON 对象”输出。JSON 的键名、嵌套结构必须与下面示例完全一致，不得新增或缺少任何键：

    [唯一允许的 JSON 结构示例]
    {json.dumps(json_schema, ensure_ascii=False, indent=2)}

    [字段含义说明]
    - "门编号"：表格中“门编号”一行对应的值。
    - "门型"：表格中“门型”一行对应的值。
    - "洞口尺寸"：表格中“洞口尺寸”一行对应的值，形如 "1490*2300"。
    - "构件尺寸"：表格中“构件尺寸”一行对应的值，形如 "1460*2300"。
    - "门框材质"：表格中“门框材质”一行的完整文本（为空则返回空字符串""）。
    - "门扇材质"：表格中“门扇材质”一行的完整文本（为空则返回空字符串""）。
    - "门槛材质"：表格中“门槛材质”一行的完整文本（为空则返回空字符串""）。
    - "防火门芯"：表格中“防火门芯”一行的完整文本（为空则返回空字符串""）。
    - "玻璃"：表格中“玻璃”一行的完整文本（为空则返回空字符串""）。
    - "门框密封条"：表格中“门框密封条”一行的完整文本（为空则返回空字符串""）。
    - "门扇密封条"：表格中“门扇密封条”一行的完整文本（为空则返回空字符串""）。
    - "五金配置组名称"：来自表头中“五金配置(XXX)”里的 XXX，只保留括号内字符串，例如“五金配置(HW-8)”→ "HW-8"、“五金配置(HW-08a)”→ "HW-08a"；如果完全没有出现，则返回空字符串""。
    - "五金配置"：对应“五金配置”表格中每一条五金明细。每一行生成一个对象，包含“名称/品牌/型号/数量”四个字段。
    - "饰面颜色"：从“饰面颜色”一行中，根据“拉门侧:…”和“推门侧:…”的描述，分别填入 "拉门侧" 和 "推门侧"。

    [抽取与清洗规则（必须执行）]
    1) 仅输出上方结构的 JSON；禁止任何解释、注释、额外键、NaN、undefined、null。若无法确定某个字段值，请使用空字符串""或空数组[]。
    2) “五金配置”表格中，“名称”若带有序号或装饰（例如：①、②、③、(1)、（1）、1.），在写入 "名称" 字段时必须去除这些序号，只保留名称文本；空行或全空项必须跳过，不出现在数组中。
    3) 在写入所有字符串字段之前，必须先反转 HTML 实体（例如：&quot; → "，&amp; → &），然后保留原有的文字内容和大小写，不进行翻译或改写。
    4) "数量" 必须是整数类型（JSON 数字，而不是字符串）；如果无法确定数量，则跳过该五金项，不要为该项生成对象。
    5) "五金配置组名称" 仅填入括号中的代码字符串，例如 "五金配置(HW-08a)" → "HW-08a"。不要包含括号或中文说明。如果未在文本中出现“五金配置(…)”结构，则将 "五金配置组名称" 设为空字符串""。
    6) "洞口尺寸" 与 "构件尺寸" 字段保留为原始的“宽*高”字符串（例如 "1490*2300"），不要增加单位、中文说明，也不要拆分为对象。
    7) "门扇材质"、"门槛材质"、"防火门芯"、"玻璃"、"门框密封条"、"门扇密封条" 等字段，直接填入对应表格行的完整文本（在反转 HTML 实体和去除首尾空白之后），不做额外解释或格式修改。
    8) "饰面颜色" 字段：从同一单元格文本中识别“拉门侧: …”和“推门侧: …”，分别填入 "拉门侧" 和 "推门侧"。如果只出现其中一侧，则另一侧使用空字符串""。
    9) 仅返回一个合法的 JSON 对象（UTF-8 编码，键和值使用双引号包裹，无尾随逗号）；禁止使用任何 Markdown 代码块标记（例如 ```json 或 ```），禁止在 JSON 前后添加多余文本。

    严格遵循以上规则，仅输出 JSON。
    """.strip()

    user_prompt: str = f"请按上面的规则，从以下markdown表格中抽取并“仅以一个JSON对象”返回：\n{markdown_text}"

    payload: dict = {
        "model": MODEL_NAME,
        "messages": [
            {"role": "system", "content": system_template},
            {"role": "user", "content": user_prompt},
        ],
        "temperature": 0,
        "do_sample": False,
        "response_format": {"type": "json_object"}
    }

    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {GLM_API_KEY}"
    }

    response = post_with_retry(CHAT_COMPLETIONS_URL, data=json.dumps(payload, ensure_ascii=False), headers=headers,
                               proxies=None)
    response.raise_for_status()

    content: str = response.json()["choices"][0]["message"]["content"]
    try:
        return json.loads(content)
    except json.JSONDecodeError as e:
        raise RuntimeError(f"GLM 返回内容非合法 JSON: {content}") from e


def pdf_process_main(pdf_path: str,
                     roi_shape: tuple[tuple[float, float], tuple[float, float]] = ((0.0, 0.0), (1.0, 1.0)),
                     dpi: int = 300, temp_png_path: Optional[str] = None) -> str:
    """
    load pdf and export png

    :param pdf_path: pdf file path
    :param roi_shape: roi shape in normalized coordinates ((x0, y0), (x1, y1))
    :param dpi: export png dpi
    :param temp_png_path: temporary png file path
    :return: cropped png base64
    """

    # load pdf and get pdf shape
    pdf = pymupdf.open(pdf_path)
    page = pdf.load_page(0)

    # calculate rotation
    rotate_deg = 0
    width, height = page.rect.width, page.rect.height
    if width < height:
        rotate_deg = -90

    # zoom and calculate pixmap
    page.set_rotation(rotation=rotate_deg)  # reset rotation

    # use roi calculate crop rect
    width, height = page.rect.width, page.rect.height
    x0, y0 = roi_shape[0]
    x1, y1 = roi_shape[1]
    print(
        f"PDF page size: width={width}, height={height}, crop rect=({x0 * width}, {y0 * height}, {x1 * width}, {y1 * height})")

    # export to matrix
    clip = pymupdf.Rect(x0 * width, y0 * height, x1 * width, y1 * height)
    clip &= page.rect  # 与 page.rect 取交集

    pix = page.get_pixmap(matrix=pymupdf.Matrix(dpi / 72, dpi / 72), alpha=False, clip=clip)

    if temp_png_path:
        pix.save(temp_png_path)

    pdf.close()

    return base64.b64encode(pix.tobytes("png")).decode('utf-8')


def multi_thread_main(cache_key: str, cache_data: cacheData):
    global CACHE_DATA_DICT
    try:
        print(f"[THREAD] Start processing: {cache_key}", flush=True)

        markdown_text: str = paddle_ocr_vl(detection=cache_data.base64)
        dumped_json: dict = big_model_completions(markdown_text=markdown_text)
        print(f"[THREAD] JSON done for {cache_key}:\n{dumped_json}", flush=True)

        with LOCK_1:
            CACHE_DATA_DICT[cache_key].markdown_text = markdown_text
            CACHE_DATA_DICT[cache_key].extracted_json = dumped_json

    except Exception as e:
        print(f"[THREAD-ERROR] key={cache_key}, error={repr(e)}", flush=True)
        with LOCK_1:
            CACHE_DATA_DICT[cache_key].extracted_json = {}


if __name__ == '__main__':
    pdf_folder: str = r"your/pdf/path"

    wb = openpyxl.Workbook()

    for dir_path in os.listdir(drawing_folder):
        full_dir_path: str = os.path.join(drawing_folder, dir_path, "pdf")
        if os.path.isdir(full_dir_path):
            output_json_path: str = os.path.join(f"./cache_json", f"{dir_path}_extracted_fhm_data.json")
            os.makedirs("./cache_json", exist_ok=True)

            # process each sub-folder
            CACHE_DATA_DICT.clear()
            print(f"Processing folder: {full_dir_path}")

            # create cache data
            os.makedirs(f"./test_img/{dir_path}", exist_ok=True)

            # preprocess pdf to cache data (pymupdf not support multi-thread)
            for file_path in os.listdir(full_dir_path):
                if file_path.endswith(".pdf") and ("FHM" in file_path.upper() or "GM" in file_path.upper()):
                    full_pdf_path = os.path.join(full_dir_path, file_path)
                    print(f"Caching PDF: {full_pdf_path}")
                    corp_base64: str = pdf_process_main(
                        pdf_path=full_pdf_path, roi_shape=((0.6, 0.55), (0.85, 1.0)),
                        temp_png_path=f"./test_img/{dir_path}/{os.path.splitext(os.path.basename(full_pdf_path))[0]}.png")
                    CACHE_DATA_DICT[file_path] = cacheData(pdf_path=full_pdf_path, base64=corp_base64)

            # 构造线程池
            pool = ThreadPoolExecutor(max_workers=min(16, (os.cpu_count() or 1) * 5))
            all_tasks = []

            for key, value in CACHE_DATA_DICT.items():
                all_tasks.append(pool.submit(multi_thread_main, key, value))

            # 等待所有任务完成
            wait(all_tasks, return_when=ALL_COMPLETED)
            pool.shutdown()

            # 输出最终JSON文件
            output_dict: dict = {}
            for key, value in CACHE_DATA_DICT.items():
                if value.extracted_json is None:
                    print(f"[WARN] {key} 在 Excel 导出时 JSON 仍为 None，跳过。", flush=True)
                    continue

                output_dict[key] = value.extracted_json

            with open(output_json_path, "w", encoding="utf-8") as f:
                json.dump(output_dict, f, ensure_ascii=False, indent=2)

            print(f"Output JSON saved to: {output_json_path}")

            # export to excel
            ws = wb.create_sheet(title=dir_path[:31])   # excel sheet title max length is 31
            # add title
            ws.append([f"防火门数据导出表 - {dir_path}"])
            ws.append(["序号", "门型", "门编号", "洞口尺寸", "构件尺寸", "门框材质", "门扇材质", "门槛材质", "防火门芯",
                       "玻璃", "门框密封条", "门扇密封条", "五金配置组", "名称", "品牌", "型号", "数量",
                       "饰面颜色-推门侧", "饰面颜色-拉门侧"])

            # merge and center the title row
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=19)
            title_cell = ws.cell(row=1, column=1)
            title_cell.alignment = styles.Alignment(horizontal='center', vertical='center')

            serial: int = 0
            for key, value in CACHE_DATA_DICT.items():
                extracted_json: dict = value.extracted_json
                hardware_list: list = extracted_json.get("五金配置", [])
                max_hardware_count: int = max(1, len(hardware_list))

                for i in range(max_hardware_count):
                    if i == 0:
                        serial += 1
                        row_data = [str(serial), extracted_json.get("门型", ""), extracted_json.get("门编号", ""),
                                    extracted_json.get("洞口尺寸", ""), extracted_json.get("构件尺寸", ""),
                                    extracted_json.get("门框材质", ""), extracted_json.get("门扇材质", ""),
                                    extracted_json.get("门槛材质", ""), extracted_json.get("防火门芯", ""),
                                    extracted_json.get("玻璃", ""), extracted_json.get("门框密封条", ""),
                                    extracted_json.get("门扇密封条", ""), extracted_json.get("五金配置组名称", ""),
                                    hardware_list[i]["名称"] if i < len(hardware_list) else "",
                                    hardware_list[i]["品牌"] if i < len(hardware_list) else "",
                                    hardware_list[i]["型号"] if i < len(hardware_list) else "",
                                    hardware_list[i]["数量"] if i < len(hardware_list) else "",
                                    extracted_json.get("饰面颜色", {}).get("推门侧", ""),
                                    extracted_json.get("饰面颜色", {}).get("拉门侧", "")]
                    else:
                        row_data = ["", "", "", "", "", "", "", "", "", "", "", "", "",
                                    hardware_list[i]["名称"] if i < len(hardware_list) else "",
                                    hardware_list[i]["品牌"] if i < len(hardware_list) else "",
                                    hardware_list[i]["型号"] if i < len(hardware_list) else "",
                                    hardware_list[i]["数量"] if i < len(hardware_list) else "", "", ""]
                    ws.append(row_data)

                # merge cells for non-hardware columns
                if max_hardware_count > 1:
                    for col_idx in range(1, 14):  # columns A to M (1 to 13)
                        ws.merge_cells(start_row=ws.max_row - max_hardware_count + 1,
                                       start_column=col_idx,
                                       end_row=ws.max_row,
                                       end_column=col_idx)
                    for col_idx in range(18, 20):  # columns R to S (18 to 19)
                        ws.merge_cells(start_row=ws.max_row - max_hardware_count + 1,
                                       start_column=col_idx,
                                       end_row=ws.max_row,
                                       end_column=col_idx)

            print(f"Excel sheet for {dir_path} done.", flush=True)

    # remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(export_xlsx)
    print(f"Exported Excel saved to: {export_xlsx}")
