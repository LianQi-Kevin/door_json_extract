""" Step.4_2 - 根据 detail_df 进行汇总和聚类，并写出到工程量清单 xlsx """
import logging
from copy import copy
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd

from config import config
from tools.logging_utils import log_set
from json_to_parquet import load_detail_parquet

EXAMPLE_XLSX = config.resources_dir / "工程量计算式模板.xlsx"


def build_package_hardware_agg_df(df: pd.DataFrame) -> pd.DataFrame:
    """ 图包级聚合表: 图包名称 + 单位 + 五金名称 + 型号 """
    return (df[df["汇总数量"].gt(0)]
            .groupby(["图包名称", "单位", "五金材料名称", "规格型号"], dropna=False, as_index=False)
            .agg(数量=("汇总数量", "sum"))
            .rename(columns={"五金材料名称": "五金名称", "规格型号": "型号"})
            .sort_values(by=["五金名称", "型号", "图包名称"], kind="stable")
            .reset_index(drop=True)[["图包名称", "单位", "五金名称", "型号", "数量"]])


def build_hardware_agg_df(df: pd.DataFrame) -> pd.DataFrame:
    """ 五金总聚合表: 单位 + 五金名称 + 型号 """
    return (df.groupby(["单位", "五金名称", "型号"], dropna=False, as_index=False)
            .agg(数量=("数量", "sum"))
            .sort_values(by=["五金名称", "型号"], kind="stable")
            .reset_index(drop=True)[["单位", "五金名称", "型号", "数量"]])


def build_package_hardware_agg_df_map(df: pd.DataFrame) -> dict[tuple[str, str], pd.DataFrame]:
    """ 按五金名称 + 型号拆分 df """
    result: dict[tuple[str, str], pd.DataFrame] = {}
    for (hardware_name, model), group_df in df.groupby(["五金名称", "型号"], sort=False, dropna=False):
        result[(str(hardware_name), str(model))] = group_df.reset_index(drop=True).copy()
    return result


def copy_row_style(ws, src_row: int, dst_row: int, max_col: int):
    """ 样式复制 """
    for col in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)

        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.protection = copy(src.protection)
            dst.number_format = src.number_format

    if src_row in ws.row_dimensions:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
        ws.row_dimensions[dst_row].hidden = ws.row_dimensions[src_row].hidden
        ws.row_dimensions[dst_row].outlineLevel = ws.row_dimensions[src_row].outlineLevel


def write_summary_sheet(wb_path: Path, df: pd.DataFrame, save_path: Optional[Path] = None):
    wb = openpyxl.load_workbook(wb_path)
    ws = wb["分部分项工程和单价措施项目清单与计价表"]

    # 获取待写入行数, 计算新增行数量
    write_pd: list[dict] = df.to_dict("records")
    ws.insert_rows(idx=7, amount=len(write_pd))
    moved_template_row = 7 + len(write_pd)
    for r in range(7, moved_template_row):
        copy_row_style(ws, moved_template_row, r, ws.max_column)

    # 写数据
    for index, (row, write_item) in enumerate(zip(ws.iter_rows(min_row=7, max_row=6 + len(write_pd)), write_pd), start=1):
        row[0].value = index
        row[1].value = f"{index:03d}"
        row[2].value = write_item["五金名称"]
        row[3].value = write_item["型号"]
        row[4].value = write_item["单位"]
        row[5].value = 0
        row[6].value = write_item["数量"]

    # 修改行高  todo: 临时使用，应当嵌入行样式复制内
    for index in range(7, moved_template_row):
        ws.row_dimensions[index].height = 30

    # 删除模板行
    ws.delete_rows(idx=moved_template_row)

    wb.save(save_path if save_path else wb_path)
    wb.close()


def write_hardware_item_sheet(wb_path: Path, df: dict[tuple[str, str], pd.DataFrame], save_path: Optional[Path] = None):
    """ 根据 <五金名称_型号> 写入独立sheet """
    wb = openpyxl.load_workbook(wb_path)

    for (name, model), df_item in df.items():
        ws = wb.copy_worksheet(from_worksheet=wb["模板_清单工程量计算表"])
        ws.title = f"{name}_{model}"

        write_list = df_item.to_dict("records")

        # 插入行 & 复制样式
        if len(write_list) > 19:
            ws.insert_rows(idx=18, amount=len(write_list) - 19)
        for r in range(8, len(write_list) + 8):
            copy_row_style(ws, 8, r, ws.max_column)

        # 写表头
        ws["F3"] = name
        ws["B4"] = write_list[0]["单位"]
        ws["B8"] = model

        # 写行
        for row, write_dict in zip(ws.iter_rows(min_row=8, max_row=8 + len(write_list)), write_list):
            row[2].value = write_dict["图包名称"].split(" ")[1]
            row[6].value = write_dict["数量"]

    # 清理模板
    del wb["模板_清单工程量计算表"]
    wb.save(save_path if save_path else wb_path)
    wb.close()


if __name__ == '__main__':
    log_set(logging.DEBUG, log_save=True, save_level=logging.WARNING, save_path=config.log_dir / "4_2_write_bill_xlsx.log")

    detail_df = load_detail_parquet()
    package_hardware_agg_df = build_package_hardware_agg_df(detail_df)

    hardware_agg_df = build_hardware_agg_df(package_hardware_agg_df)
    result_dict = build_package_hardware_agg_df_map(package_hardware_agg_df)

    xlsx_export_path: Path = Path("./五金工程量清单.xlsx")
    write_summary_sheet(EXAMPLE_XLSX, hardware_agg_df, save_path=xlsx_export_path)
    write_hardware_item_sheet(xlsx_export_path, result_dict)
